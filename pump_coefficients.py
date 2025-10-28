import math
import re
import threading
import uuid
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook

import equipment_selection


class PumpCoefficientError(Exception):
    """Error genérico para operaciones de coeficientes."""


class PumpCoefficientValidationError(PumpCoefficientError):
    def __init__(self, errors: Dict[str, str]):
        super().__init__('Errores de validación en los coeficientes de bomba.')
        self.errors = errors


_LOCK = threading.Lock()
_ID_HEADER = 'id'

_LABEL_OVERRIDES = {
    'Bomba': 'Modelo',
    'nom': 'Caudal nominal',
    'min': 'Caudal mínimo',
    'max': 'Caudal máximo',
    'maxmax': 'Caudal máximo permitido',
    'rpm': 'RPM',
}


def _ensure_catalog_loaded() -> None:
    if equipment_selection.PUMP_CATALOG is None:
        equipment_selection.load_catalogs()


def _open_sheet():
    _ensure_catalog_loaded()
    path = equipment_selection.get_catalog_excel_path()
    sheet_name = equipment_selection.get_pump_sheet_name()

    if not sheet_name:
        raise PumpCoefficientError('No se pudo determinar la hoja de bombas en el archivo de catálogos.')

    workbook = load_workbook(path)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise PumpCoefficientError(f"La hoja '{sheet_name}' no existe en el archivo de catálogo.")

    return workbook, workbook[sheet_name], path, sheet_name


def _ensure_id_column(worksheet) -> Tuple[List[str], int, bool]:
    headers: List[str] = []
    for cell in worksheet[1]:
        headers.append(cell.value if cell.value is not None else '')

    id_index = None
    for idx, header in enumerate(headers):
        if isinstance(header, str) and header.strip().lower() == _ID_HEADER:
            id_index = idx
            headers[idx] = _ID_HEADER
            break

    changed = False
    if id_index is None:
        id_index = len(headers)
        worksheet.cell(row=1, column=id_index + 1, value=_ID_HEADER)
        headers.append(_ID_HEADER)
        changed = True

    max_row = worksheet.max_row or 1
    for row_idx in range(2, max_row + 1):
        row_cells = worksheet[row_idx]
        if len(row_cells) < id_index + 1:
            worksheet.cell(row=row_idx, column=id_index + 1, value=str(uuid.uuid4()))
            changed = True
            continue

        id_cell = row_cells[id_index]
        row_is_empty = all((cell.value is None or str(cell.value).strip() == '') for cell in row_cells[:-1])
        if row_is_empty:
            continue

        if not id_cell.value or str(id_cell.value).strip() == '':
            id_cell.value = str(uuid.uuid4())
            changed = True

    return headers, id_index, changed


def _detect_column_types(df: pd.DataFrame) -> Dict[str, str]:
    numeric_hints = set(
        equipment_selection.COLS_PUMP_HEAD
        + equipment_selection.COLS_PUMP_BHP
        + equipment_selection.COLS_PUMP_EFF
    )
    numeric_hints.update(
        col
        for col in (
            equipment_selection.COL_PUMP_MIN_Q,
            equipment_selection.COL_PUMP_MAX_Q,
        )
        if col
    )

    column_types: Dict[str, str] = {}
    for column in df.columns:
        if column == _ID_HEADER:
            column_types[column] = 'string'
            continue

        if column in numeric_hints:
            column_types[column] = 'number'
            continue

        series = df[column]
        if pd.api.types.is_numeric_dtype(series.dropna()) and series.dropna().shape[0] > 0:
            column_types[column] = 'number'
        else:
            column_types[column] = 'string'

    return column_types


def _determine_required_columns(columns: List[str]) -> List[str]:
    required = set()
    primary_id = equipment_selection.COL_PUMP_ID
    if primary_id in columns:
        required.add(primary_id)

    for col in (equipment_selection.COL_PUMP_MIN_Q, equipment_selection.COL_PUMP_MAX_Q):
        if col in columns:
            required.add(col)

    for col in equipment_selection.COLS_PUMP_HEAD:
        if col in columns:
            required.add(col)

    for col in equipment_selection.COLS_PUMP_BHP:
        if col in columns:
            required.add(col)

    return list(required)


def _read_table() -> Tuple[pd.DataFrame, Dict[str, str], List[str]]:
    _ensure_catalog_loaded()
    path = equipment_selection.get_catalog_excel_path()
    sheet_name = equipment_selection.get_pump_sheet_name()

    df = pd.read_excel(path, sheet_name=sheet_name)
    if _ID_HEADER not in df.columns:
        df[_ID_HEADER] = None

    df = df.where(pd.notna(df), None)
    column_types = _detect_column_types(df)
    required_columns = _determine_required_columns(list(df.columns))
    return df, column_types, required_columns


def _normalize_row(
    payload: Dict[str, Any],
    columns: List[str],
    column_types: Dict[str, str],
    required_columns: List[str],
    *,
    is_create: bool
) -> Dict[str, Any]:
    errors: Dict[str, str] = {}
    normalized: Dict[str, Any] = {}

    allowed_keys = set(columns)
    allowed_keys.discard(_ID_HEADER)

    extra_keys = set(payload.keys()) - allowed_keys - {_ID_HEADER}
    if extra_keys:
        errors['_extra'] = f"Columnas no reconocidas: {', '.join(sorted(extra_keys))}"

    for column in columns:
        if column == _ID_HEADER:
            continue

        raw_value = payload.get(column)

        if (column in required_columns) and (raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == '')):
            errors[column] = 'Campo requerido'
            continue

        if raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == ''):
            normalized[column] = None
            continue

        if column_types.get(column) == 'number':
            try:
                normalized[column] = float(raw_value)
            except (TypeError, ValueError):
                errors[column] = 'Debe ser numérico'
        else:
            normalized[column] = str(raw_value).strip()

    if errors:
        raise PumpCoefficientValidationError(errors)

    return normalized


def _serialise_rows(df: pd.DataFrame, column_types: Dict[str, str]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for record in df.to_dict(orient='records'):
        row: Dict[str, Any] = {}
        empty_content = True
        for column, value in record.items():
            if column == _ID_HEADER:
                row[column] = str(value) if value is not None else None
                continue

            if value is None:
                row[column] = None
            elif column_types.get(column) == 'number':
                try:
                    number_value = float(value)
                except (TypeError, ValueError):
                    number_value = None

                if number_value is None or math.isnan(number_value):
                    row[column] = None
                else:
                    row[column] = number_value
            else:
                row[column] = str(value)

            if row[column] not in (None, ''):
                empty_content = False

        if row.get(_ID_HEADER) is None:
            row[_ID_HEADER] = str(uuid.uuid4())

        if empty_content:
            continue

        rows.append(row)

    return rows


def _format_column_label(column: str) -> str:
    if not column:
        return ''

    override = _LABEL_OVERRIDES.get(column)
    if override:
        return override

    label = column.replace('_', ' ')
    label = re.sub(r'([a-z])([A-Z])', r'\1 \2', label)
    label = re.sub(r'(\d+)', r' \1', label)
    label = re.sub(r'\s+', ' ', label).strip()

    if not label:
        return column

    return label[0].upper() + label[1:]


def list_pump_coefficients() -> Dict[str, Any]:
    with _LOCK:
        workbook, worksheet, path, sheet_name = _open_sheet()
        headers, _, changed = _ensure_id_column(worksheet)
        if changed:
            workbook.save(path)
        workbook.close()

        df, column_types, required_columns = _read_table()
        rows = _serialise_rows(df, column_types)

    display_column = equipment_selection.COL_PUMP_ID

    return {
        'success': True,
        'columns': [
            {
                'key': column,
                'type': column_types.get(column, 'string'),
                'required': column in required_columns,
                'read_only': column == _ID_HEADER,
                'label': _format_column_label(column),
            }
            for column in df.columns
        ],
        'id_column': _ID_HEADER,
        'display_column': display_column,
        'data': rows,
    }


def _find_row_index_by_id(worksheet, headers: List[str], target_id: str) -> int:
    try:
        id_idx = headers.index(_ID_HEADER)
    except ValueError as exc:
        raise PumpCoefficientError('La hoja de bombas no tiene columna id.') from exc

    max_row = worksheet.max_row or 1
    for row_idx in range(2, max_row + 1):
        cell_value = worksheet.cell(row=row_idx, column=id_idx + 1).value
        if cell_value is None:
            continue
        if str(cell_value).strip() == target_id:
            return row_idx
    raise PumpCoefficientError(f'No se encontró la bomba con id {target_id}.')


def create_pump(payload: Dict[str, Any]) -> Dict[str, Any]:
    with _LOCK:
        workbook, worksheet, path, _ = _open_sheet()
        headers, _, changed = _ensure_id_column(worksheet)
        if changed:
            workbook.save(path)
        workbook.close()

        df, column_types, required_columns = _read_table()
        normalized = _normalize_row(payload, list(df.columns), column_types, required_columns, is_create=True)

        new_id = str(uuid.uuid4())
        workbook, worksheet, path, _ = _open_sheet()
        headers, _, changed = _ensure_id_column(worksheet)
        if changed:
            workbook.save(path)

        new_row_values = []
        for column in df.columns:
            if column == _ID_HEADER:
                new_row_values.append(new_id)
            else:
                new_row_values.append(normalized.get(column))

        worksheet.append(new_row_values)
        workbook.save(path)
        workbook.close()

        equipment_selection.refresh_catalogs()

        df, column_types, _ = _read_table()
        df = df[df[_ID_HEADER] == new_id]
        rows = _serialise_rows(df, column_types)
        return rows[0] if rows else {}


def update_pump(pump_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    with _LOCK:
        workbook, worksheet, path, _ = _open_sheet()
        headers, _, changed = _ensure_id_column(worksheet)
        if changed:
            workbook.save(path)
        workbook.close()

        df, column_types, required_columns = _read_table()
        normalized = _normalize_row(payload, list(df.columns), column_types, required_columns, is_create=False)

        workbook, worksheet, path, _ = _open_sheet()
        headers, _, changed = _ensure_id_column(worksheet)
        if changed:
            workbook.save(path)

        row_idx = _find_row_index_by_id(worksheet, headers, str(pump_id))

        column_index_map = {header: idx for idx, header in enumerate(headers)}
        for column, value in normalized.items():
            if column not in column_index_map:
                continue
            excel_col_idx = column_index_map[column] + 1
            worksheet.cell(row=row_idx, column=excel_col_idx, value=value)

        workbook.save(path)
        workbook.close()

        equipment_selection.refresh_catalogs()

        df, column_types, _ = _read_table()
        df = df[df[_ID_HEADER] == str(pump_id)]
        rows = _serialise_rows(df, column_types)
        return rows[0] if rows else {}


def delete_pump(pump_id: str) -> None:
    with _LOCK:
        workbook, worksheet, path, _ = _open_sheet()
        headers, _, changed = _ensure_id_column(worksheet)
        if changed:
            workbook.save(path)
        workbook.close()

        workbook, worksheet, path, _ = _open_sheet()
        headers, _, changed = _ensure_id_column(worksheet)
        if changed:
            workbook.save(path)

        target_row = _find_row_index_by_id(worksheet, headers, str(pump_id))
        worksheet.delete_rows(target_row, 1)
        workbook.save(path)
        workbook.close()

        equipment_selection.refresh_catalogs()
